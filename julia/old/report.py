import os
import warnings

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings("ignore")

BASE_DIR = "."
VALIDATION_REPORT = os.path.join(BASE_DIR, "kenza_excel_validation_report.csv")
EXPECTED_CSV = os.path.join(BASE_DIR, "kenza_excel_validation_expected.csv")
OUTPUT_XLSX = os.path.join(BASE_DIR, "kenza_validation_report.xlsx")


def load_data():
    """Charge le rapport de validation et les valeurs attendues Excel."""
    report = pd.read_csv(VALIDATION_REPORT, sep=";", decimal=",")
    expected = pd.read_csv(EXPECTED_CSV, sep=",")
    return report, expected


def prepare_model_data(report, expected, model_name, excel_col):
    model_report = report[report["model"] == model_name].copy()
    merged = expected[["year", "actual_passengers"]].merge(
        model_report[["year", "julia_prediction"]],
        on="year",
        how="left",
    )
    merged["excel_prediction"] = expected[excel_col] if excel_col in expected.columns else None
    return merged.sort_values("year").reset_index(drop=True)


def create_chart(ws, model_name, data_start_row=1, data_end_row=None, anchor="G2"):
    """Cree un graphique Excel natif pour Historique, Julia et Excel."""
    if data_end_row is None:
        data_end_row = ws.max_row

    chart = LineChart()
    chart.title = f"Modele : {model_name}"
    chart.y_axis.title = "Passagers (millions)"
    chart.x_axis.title = "Annee"
    chart.height = 10
    chart.width = 20

    data = Reference(ws, min_col=2, max_col=4, min_row=data_start_row, max_row=data_end_row)
    categories = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=data_end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    styles = [("000000", None), ("1F77B4", None), ("D62728", "dash")]
    for series, (color, dash_style) in zip(chart.series, styles):
        series.graphicalProperties.line.solidFill = color
        series.graphicalProperties.line.width = 25000
        if dash_style:
            series.graphicalProperties.line.dashStyle = dash_style
        series.marker.symbol = "circle"
        series.marker.size = 5

    ws.add_chart(chart, anchor)
    return chart


def create_abs_error_chart(ws, model_name, data_start_row=1, data_end_row=None, anchor="G22"):
    """Cree un graphique Excel natif avec la colonne abs_error."""
    if data_end_row is None:
        data_end_row = ws.max_row

    chart = BarChart()
    chart.type = "col"
    chart.title = f"Erreur absolue : {model_name}"
    chart.y_axis.title = "abs_error"
    chart.x_axis.title = "Annee"
    chart.height = 9
    chart.width = 20

    data = Reference(ws, min_col=5, min_row=data_start_row, max_row=data_end_row)
    categories = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=data_end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = "F28E2B"
        chart.series[0].graphicalProperties.line.solidFill = "F28E2B"

    ws.add_chart(chart, anchor)
    return chart


def compute_metrics(df):
    mask = df["julia_prediction"].notna() & df["excel_prediction"].notna()
    if not mask.any():
        return None
    julia = df.loc[mask, "julia_prediction"]
    excel = df.loc[mask, "excel_prediction"]
    errors = julia - excel
    return {
        "MAE": errors.abs().mean(),
        "RMSE": (errors**2).mean() ** 0.5,
        "MAPE": (errors.abs() / excel.abs()).mean() * 100,
        "N": len(julia),
    }


def append_metrics(ws, df):
    metrics = compute_metrics(df)
    if not metrics:
        return
    row_metrics = ws.max_row + 2
    ws.cell(row=row_metrics, column=1, value="Metriques (Julia vs Excel)")
    ws.cell(row=row_metrics + 1, column=1, value="MAE")
    ws.cell(row=row_metrics + 1, column=2, value=metrics["MAE"])
    ws.cell(row=row_metrics + 2, column=1, value="RMSE")
    ws.cell(row=row_metrics + 2, column=2, value=metrics["RMSE"])
    ws.cell(row=row_metrics + 3, column=1, value="MAPE (%)")
    ws.cell(row=row_metrics + 3, column=2, value=metrics["MAPE"])
    ws.cell(row=row_metrics + 4, column=1, value="N (annees)")
    ws.cell(row=row_metrics + 4, column=2, value=metrics["N"])


def main():
    print("Chargement des donnees...")
    report, expected = load_data()

    model_map = {
        "kenza": "excel_full_forecast",
        "kenza_simplifie": "excel_simplified_forecast",
        "kenza_simplifie_combine": "excel_simplified_forecast",
        "kenza_simplifie_indexe": "excel_indexed_forecast",
        "kenza_indexed": "excel_indexed_forecast",
    }

    wb = Workbook()
    wb.remove(wb.active)

    ws_data = wb.create_sheet("Donnees")
    for row in dataframe_to_rows(report, index=False, header=True):
        ws_data.append(row)

    for model_name, excel_col in model_map.items():
        print(f"Traitement du modele : {model_name}")
        df = prepare_model_data(report, expected, model_name, excel_col)
        if df.empty:
            continue

        ws = wb.create_sheet(model_name[:31])
        ws.append(["Annee", "Historique", "Julia", "Excel", "abs_error"])
        for _, row in df.iterrows():
            abs_error = None
            if pd.notna(row["julia_prediction"]) and pd.notna(row["excel_prediction"]):
                abs_error = abs(row["julia_prediction"] - row["excel_prediction"])
            ws.append([
                row["year"],
                row["actual_passengers"],
                row["julia_prediction"],
                row["excel_prediction"],
                abs_error,
            ])

        table_end_row = ws.max_row
        create_chart(ws, model_name, data_end_row=table_end_row, anchor="G2")
        create_abs_error_chart(ws, model_name, data_end_row=table_end_row, anchor="G22")
        append_metrics(ws, df)

        for col in ["A", "B", "C", "D", "E"]:
            ws.column_dimensions[col].width = 15

    wb.save(OUTPUT_XLSX)
    print(f"Fichier Excel genere : {OUTPUT_XLSX}")

    ws.column_dimensions[col].width = 15

    # Sauvegarder
    wb.save(OUTPUT_XLSX)
    print(f"Fichier Excel généré : {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()
